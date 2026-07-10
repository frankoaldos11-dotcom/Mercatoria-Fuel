mport datetime
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, session
from database import conectar
from utils.constants import TIPOS_COMBUSTIBLE, TIPOS_COMBUSTIBLE_LABELS, ROLES_ADMIN_PM
from utils.auth import requiere_login, requiere_staff

unidades_bp = Blueprint("unidades", __name__, url_prefix="/unidades")

_COMBUSTIBLE_NORM = {
    "diesel":            "diesel",
    "diésel":           "diesel",
    "gasolina regular":  "gasolina_regular",
    "regular":           "gasolina_regular",
    "gasolina_regular":  "gasolina_regular",
    "gasolina especial": "gasolina_especial",
    "especial":          "gasolina_especial",
    "gasolina_especial": "gasolina_especial",
}

_COLUMN_MAP = {
    # Vehículo
    "chapa":              "chapa",
    "matrícula":         "chapa",
    "matricula":          "chapa",
    "marca":              "marca",
    "modelo":             "modelo",
    "año":               "anio",
    "ano":                "anio",
    "anio":               "anio",
    "tipo_combustible":   "tipo_combustible",
    "tipo combustible":   "tipo_combustible",
    "combustible":        "tipo_combustible",
    "cuota_mensual_l":    "cuota_mensual_l",
    "cuota mensual":      "cuota_mensual_l",
    "cuota mensual (l)":  "cuota_mensual_l",
    "cuota (l)":          "cuota_mensual_l",
    "color":              "color",
    "cliente":            "cliente_codigo",
    "cliente_codigo":     "cliente_codigo",
    "codigo cliente":     "cliente_codigo",
    "código cliente":    "cliente_codigo",
    # Chofer / Conductor
    "chofer":             "chofer_nombre",
    "conductor":          "chofer_nombre",
    "chofer_nombre":      "chofer_nombre",
    "nombre chofer":      "chofer_nombre",
    "chofer_ci":          "chofer_ci",
    "ci chofer":          "chofer_ci",
    "ci":                 "chofer_ci",
    "carnet":             "chofer_ci",
    "carnet de identidad":"chofer_ci",
    "licencia_numero":    "licencia_numero",
    "licencia numero":    "licencia_numero",
    "n° licencia":       "licencia_numero",
    "licencia":           "licencia_numero",
    "licencia_vencimiento":"licencia_vencimiento",
    "vencimiento":        "licencia_vencimiento",
    "vencimiento licencia":"licencia_vencimiento",
    "vence":              "licencia_vencimiento",
    "telefono":           "telefono",
    "teléfono":          "telefono",
    "observaciones":      "observaciones",
}


def _requiere_admin_pm():
    return session.get("rol") not in ROLES_ADMIN_PM


def _registrar_auditoria(usuario_id, accion, tabla, registro_id, valor_anterior=None, valor_nuevo=None):
    try:
        conn = conectar()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO auditoria
                (usuario_id, accion, tabla_afectada, registro_id, valor_anterior, valor_nuevo, ip, user_agent)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            usuario_id, accion, tabla, registro_id,
            str(valor_anterior) if valor_anterior else None,
            str(valor_nuevo) if valor_nuevo else None,
            request.remote_addr,
            request.headers.get("User-Agent", "")[:512],
        ))
        conn.commit()
        conn.close()
    except Exception:
        from flask import current_app
        current_app.logger.exception("Error registrando auditoría")


def _parse_fecha(val):
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _upsert_chofer(cur, cliente_id, ci, nombre, licencia_numero, licencia_vencimiento,
                   telefono, observaciones, existing_chofer_id=None):
    """
    Devuelve el ID del chofer a vincular.
    Lógica:
      - Si el CI corresponde a un chofer existente: actualiza sus datos y devuelve su id.
      - Si no existe por CI: si el vehículo ya tenía un chofer vinculado, actualiza ese
        registro con el nuevo CI. Si no: crea uno nuevo.
    """
    if not ci or not nombre:
        return None

    cur.execute("SELECT id FROM choferes WHERE ci = ?", (ci,))
    row = cur.fetchone()

    if row:
        chofer_id = row["id"]
        cur.execute("""
            UPDATE choferes
            SET cliente_id = ?, nombre = ?, licencia_numero = ?,
                licencia_vencimiento = ?, telefono = ?, observaciones = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (cliente_id, nombre, licencia_numero or None, licencia_vencimiento or None,
              telefono or None, observaciones or None, chofer_id))
    elif existing_chofer_id:
        chofer_id = existing_chofer_id
        cur.execute("""
            UPDATE choferes
            SET cliente_id = ?, ci = ?, nombre = ?, licencia_numero = ?,
                licencia_vencimiento = ?, telefono = ?, observaciones = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (cliente_id, ci, nombre, licencia_numero or None, licencia_vencimiento or None,
              telefono or None, observaciones or None, chofer_id))
    else:
        cur.execute("""
            INSERT INTO choferes
                (cliente_id, ci, nombre, licencia_numero, licencia_vencimiento,
                 telefono, observaciones)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (cliente_id, ci, nombre, licencia_numero or None, licencia_vencimiento or None,
              telefono or None, observaciones or None))
        chofer_id = cur.lastrowid

    return chofer_id


def _importar_excel(archivo):
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(archivo.read()), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(cell.value or "").strip().lower() for cell in header_row]

    col_indices = {}
    for i, h in enumerate(headers):
        mapped = _COLUMN_MAP.get(h)
        if mapped and mapped not in col_indices:
            col_indices[mapped] = i

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT id, codigo FROM clientes WHERE activo = 1")
    clientes_map = {row["codigo"]: row["id"] for row in cur.fetchall()}

    importados = 0
    fallos = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        def get(col):
            idx = col_indices.get(col)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        chapa = get("chapa")
        if not chapa:
            fallos.append(f"Fila {row_num}: chapa vacía — fila omitida")
            continue
        chapa = chapa.upper()

        cliente_codigo = get("cliente_codigo")
        cliente_id = clientes_map.get(cliente_codigo) if cliente_codigo else None
        if not cliente_id:
            fallos.append(f"Fila {row_num} ({chapa}): cliente '{cliente_codigo}' no encontrado — fila omitida")
            continue

        combustible_raw = (get("tipo_combustible") or "").strip().lower()
        tipo_combustible = _COMBUSTIBLE_NORM.get(combustible_raw)
        if not tipo_combustible:
            fallos.append(f"Fila {row_num} ({chapa}): tipo de combustible inválido '{combustible_raw}' — fila omitida")
            continue

        marca = get("marca")
        modelo = get("modelo")
        anio_raw = get("anio")
        try:
            anio = int(float(anio_raw)) if anio_raw else None
        except (ValueError, TypeError):
            anio = None

        cuota_raw = get("cuota_mensual_l")
        try:
            cuota = float(cuota_raw) if cuota_raw else None
        except (ValueError, TypeError):
            cuota = None

        color = get("color")
        observaciones = get("observaciones")

        chofer_nombre = get("chofer_nombre")
        chofer_ci = get("chofer_ci")
        licencia_numero = get("licencia_numero")
        venc_raw = row[col_indices["licencia_vencimiento"]] if "licencia_vencimiento" in col_indices else None
        licencia_vencimiento = _parse_fecha(venc_raw)
        telefono = get("telefono")

        try:
            cur.execute("SELECT id, chofer_id FROM vehiculos WHERE chapa = ?", (chapa,))
            existente = cur.fetchone()

            chofer_id = None
            if chofer_ci and chofer_nombre:
                existing_ch_id = existente["chofer_id"] if existente else None
                chofer_id = _upsert_chofer(
                    cur, cliente_id, chofer_ci, chofer_nombre,
                    licencia_numero, licencia_vencimiento, telefono, None,
                    existing_chofer_id=existing_ch_id
                )

            if existente:
                cur.execute("""
                    UPDATE vehiculos
                    SET cliente_id = ?, marca = ?, modelo = ?, anio = ?,
                        tipo_combustible = ?, cuota_mensual_l = ?, color = ?,
                        observaciones = ?, chofer_id = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE chapa = ?
                """, (cliente_id, marca, modelo, anio, tipo_combustible, cuota, color,
                      observaciones, chofer_id, chapa))
            else:
                cur.execute("""
                    INSERT INTO vehiculos
                        (cliente_id, chapa, marca, modelo, anio, tipo_combustible,
                         cuota_mensual_l, color, observaciones, chofer_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (cliente_id, chapa, marca, modelo, anio, tipo_combustible,
                      cuota, color, observaciones, chofer_id))
            importados += 1
        except Exception as exc:
            fallos.append(f"Fila {row_num} ({chapa}): error al guardar — {exc}")

    conn.commit()
    conn.close()
    return importados, fallos


@unidades_bp.route("/")
def listado():
    redir = requiere_staff()
    if redir:
        return redir

    buscar = request.args.get("buscar", "").strip()
    filtro_cliente = request.args.get("cliente_id", "").strip()
    filtro_combustible = request.args.get("combustible", "").strip()
    filtro_estado = request.args.get("estado", "").strip()

    condiciones = []
    params = []

    if buscar:
        condiciones.append("""(
            v.chapa LIKE ? OR v.marca LIKE ? OR v.modelo LIKE ?
            OR ch.nombre LIKE ? OR ch.ci LIKE ?
            OR cli.nombre LIKE ?
        )""")
        like = f"%{buscar}%"
        params.extend([like, like, like, like, like, like])
    if filtro_cliente:
        condiciones.append("v.cliente_id = ?")
        params.append(filtro_cliente)
    if filtro_combustible:
        condiciones.append("v.tipo_combustible = ?")
        params.append(filtro_combustible)
    if filtro_estado:
        condiciones.append("v.estado = ?")
        params.append(filtro_estado)

    where = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""

    conn = conectar()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT v.id, v.cliente_id, v.chapa, v.marca, v.modelo, v.anio,
               v.tipo_combustible, v.cuota_mensual_l, v.estado, v.chofer_id,
               cli.nombre AS cliente_nombre, cli.codigo AS cliente_codigo,
               ch.nombre AS chofer_nombre, ch.ci AS chofer_ci,
               ch.licencia_numero, ch.licencia_vencimiento, ch.telefono,
               ch.estado AS chofer_estado
        FROM vehiculos v
        JOIN clientes cli ON cli.id = v.cliente_id
        LEFT JOIN choferes ch ON ch.id = v.chofer_id
        {where}
        ORDER BY cli.nombre ASC, v.chapa ASC
    """, params)
    lista = cur.fetchall()

    cur.execute("SELECT id, nombre, codigo FROM clientes WHERE activo = 1 ORDER BY nombre ASC")
    clientes = cur.fetchall()
    conn.close()

    hoy = datetime.date.today().isoformat()
    limite_30 = (datetime.date.today() + datetime.timedelta(days=30)).isoformat()

    return render_template(
        "unidades/listado.html",
        lista=lista,
        clientes=clientes,
        buscar=buscar,
        filtro_cliente=filtro_cliente,
        filtro_combustible=filtro_combustible,
        filtro_estado=filtro_estado,
        tipos_combustible=TIPOS_COMBUSTIBLE,
        combustible_labels=TIPOS_COMBUSTIBLE_LABELS,
        hoy=hoy,
        limite_30=limite_30,
    )


@unidades_bp.route("/crear", methods=["GET", "POST"])
def crear():
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/unidades?access_error=Solo+Admin+y+PM+pueden+crear+unidades")

    error = None

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, codigo FROM clientes WHERE activo = 1 ORDER BY nombre ASC")
    clientes = cur.fetchall()
    conn.close()

    if request.method == "POST":
        # — Vehículo —
        cliente_id   = request.form.get("cliente_id", "").strip()
        chapa        = request.form.get("chapa", "").strip().upper()
        marca        = request.form.get("marca", "").strip()
        modelo       = request.form.get("modelo", "").strip()
        anio_str     = request.form.get("anio", "").strip()
        tipo_comb    = request.form.get("tipo_combustible", "").strip()
        cuota_str    = request.form.get("cuota_mensual_l", "").strip()
        color        = request.form.get("color", "").strip()
        observaciones= request.form.get("observaciones", "").strip()
        estado       = request.form.get("estado", "activo").strip()
        # — Chofer —
        chofer_nombre = request.form.get("chofer_nombre", "").strip()
        chofer_ci     = request.form.get("chofer_ci", "").strip()
        licencia_num  = request.form.get("licencia_numero", "").strip()
        licencia_venc = request.form.get("licencia_vencimiento", "").strip() or None
        telefono      = request.form.get("telefono", "").strip()

        if not chapa:
            error = "La chapa del vehículo es obligatoria."
        elif not cliente_id:
            error = "Debe seleccionar un cliente."
        elif tipo_comb not in TIPOS_COMBUSTIBLE:
            error = "Tipo de combustible no válido."
        elif bool(chofer_nombre) != bool(chofer_ci):
            error = "Si registra un chofer, debe ingresar tanto el nombre como el C.I."
        else:
            try:
                anio = int(anio_str) if anio_str else None
            except ValueError:
                anio = None
            try:
                cuota = float(cuota_str) if cuota_str else None
            except ValueError:
                cuota = None

            conn = conectar()
            cur = conn.cursor()
            cur.execute("SELECT id FROM vehiculos WHERE chapa = ?", (chapa,))
            if cur.fetchone():
                error = f"Ya existe una unidad con la chapa {chapa}."
                conn.close()
            else:
                chofer_id = _upsert_chofer(
                    cur, int(cliente_id), chofer_ci, chofer_nombre,
                    licencia_num, licencia_venc, telefono, None
                ) if chofer_ci and chofer_nombre else None

                cur.execute("""
                    INSERT INTO vehiculos
                        (cliente_id, chapa, marca, modelo, anio, tipo_combustible,
                         cuota_mensual_l, color, observaciones, estado, chofer_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (cliente_id, chapa, marca, modelo, anio, tipo_comb,
                      cuota, color or None, observaciones or None, estado, chofer_id))
                nuevo_id = cur.lastrowid
                conn.commit()
                conn.close()
                _registrar_auditoria(
                    session.get("user_id"), "Creó unidad autorizada", "vehiculos", nuevo_id,
                    valor_nuevo={"chapa": chapa, "cliente_id": cliente_id,
                                 "tipo_combustible": tipo_comb, "chofer_ci": chofer_ci}
                )
                return redirect("/unidades?ok=1")

    return render_template(
        "unidades/crear.html",
        error=error,
        clientes=clientes,
        tipos_combustible=TIPOS_COMBUSTIBLE,
        combustible_labels=TIPOS_COMBUSTIBLE_LABELS,
        cliente_preseleccionado=request.args.get("cliente_id", ""),
    )


@unidades_bp.route("/<int:id>/editar", methods=["GET", "POST"])
def editar(id):
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/unidades?access_error=Solo+Admin+y+PM+pueden+editar+unidades")

    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT v.*, cli.nombre AS cliente_nombre,
               ch.nombre AS chofer_nombre, ch.ci AS chofer_ci,
               ch.licencia_numero, ch.licencia_vencimiento, ch.telefono
        FROM vehiculos v
        JOIN clientes cli ON cli.id = v.cliente_id
        LEFT JOIN choferes ch ON ch.id = v.chofer_id
        WHERE v.id = ?
    """, (id,))
    unidad = cur.fetchone()

    if not unidad:
        conn.close()
        return redirect("/unidades")

    cur.execute("SELECT id, nombre, codigo FROM clientes WHERE activo = 1 ORDER BY nombre ASC")
    clientes = cur.fetchall()

    error = None

    if request.method == "POST":
        cliente_id    = request.form.get("cliente_id", "").strip()
        chapa         = request.form.get("chapa", "").strip().upper()
        marca         = request.form.get("marca", "").strip()
        modelo        = request.form.get("modelo", "").strip()
        anio_str      = request.form.get("anio", "").strip()
        tipo_comb     = request.form.get("tipo_combustible", "").strip()
        cuota_str     = request.form.get("cuota_mensual_l", "").strip()
        color         = request.form.get("color", "").strip()
        observaciones = request.form.get("observaciones", "").strip()
        estado        = request.form.get("estado", "activo").strip()
        chofer_nombre = request.form.get("chofer_nombre", "").strip()
        chofer_ci     = request.form.get("chofer_ci", "").strip()
        licencia_num  = request.form.get("licencia_numero", "").strip()
        licencia_venc = request.form.get("licencia_vencimiento", "").strip() or None
        telefono      = request.form.get("telefono", "").strip()

        if not chapa:
            error = "La chapa del vehículo es obligatoria."
        elif not cliente_id:
            error = "Debe seleccionar un cliente."
        elif tipo_comb not in TIPOS_COMBUSTIBLE:
            error = "Tipo de combustible no válido."
        elif bool(chofer_nombre) != bool(chofer_ci):
            error = "Si registra un chofer, debe ingresar tanto el nombre como el C.I."
        else:
            cur.execute("SELECT id FROM vehiculos WHERE chapa = ? AND id != ?", (chapa, id))
            if cur.fetchone():
                error = f"Ya existe otra unidad con la chapa {chapa}."
            else:
                try:
                    anio = int(anio_str) if anio_str else None
                except ValueError:
                    anio = None
                try:
                    cuota = float(cuota_str) if cuota_str else None
                except ValueError:
                    cuota = None

                existing_ch_id = unidad["chofer_id"]
                if chofer_ci and chofer_nombre:
                    chofer_id = _upsert_chofer(
                        cur, int(cliente_id), chofer_ci, chofer_nombre,
                        licencia_num, licencia_venc, telefono, None,
                        existing_chofer_id=existing_ch_id
                    )
                else:
                    chofer_id = None

                anterior = dict(unidad)
                cur.execute("""
                    UPDATE vehiculos
                    SET cliente_id = ?, chapa = ?, marca = ?, modelo = ?, anio = ?,
                        tipo_combustible = ?, cuota_mensual_l = ?, color = ?,
                        observaciones = ?, estado = ?, chofer_id = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (cliente_id, chapa, marca, modelo, anio, tipo_comb,
                      cuota, color or None, observaciones or None, estado, chofer_id, id))
                conn.commit()
                conn.close()
                _registrar_auditoria(
                    session.get("user_id"), "Editó unidad autorizada", "vehiculos", id,
                    valor_anterior=anterior,
                    valor_nuevo={"chapa": chapa, "tipo_combustible": tipo_comb,
                                 "estado": estado, "chofer_ci": chofer_ci}
                )
                return redirect("/unidades?ok=1")

    conn.close()
    return render_template(
        "unidades/editar.html",
        unidad=unidad,
        clientes=clientes,
        error=error,
        tipos_combustible=TIPOS_COMBUSTIBLE,
        combustible_labels=TIPOS_COMBUSTIBLE_LABELS,
    )


@unidades_bp.route("/<int:id>/toggle", methods=["POST"])
def toggle_estado(id):
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/unidades?access_error=Sin+permisos")

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT estado FROM vehiculos WHERE id = ?", (id,))
    row = cur.fetchone()
    if row:
        nuevo_estado = "inactivo" if row["estado"] == "activo" else "activo"
        cur.execute(
            "UPDATE vehiculos SET estado = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (nuevo_estado, id)
        )
        conn.commit()
        _registrar_auditoria(
            session.get("user_id"),
            f"Cambió estado unidad a {nuevo_estado}", "vehiculos", id,
            valor_anterior={"estado": row["estado"]},
            valor_nuevo={"estado": nuevo_estado}
        )
    conn.close()
    return redirect("/unidades")


@unidades_bp.route("/importar", methods=["GET", "POST"])
def importar():
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/unidades?access_error=Solo+Admin+y+PM+pueden+importar+unidades")

    resultado = None

    if request.method == "POST":
        archivo = request.files.get("archivo")
        if not archivo or not archivo.filename.endswith(".xlsx"):
            resultado = {"error": "Debe seleccionar un archivo .xlsx válido."}
        else:
            try:
                importados, fallos = _importar_excel(archivo)
                resultado = {"importados": importados, "fallos": fallos}
            except Exception as exc:
                resultado = {"error": f"Error al procesar el archivo: {exc}"}

    return render_template("unidades/importar.html", resultado=resultado)
