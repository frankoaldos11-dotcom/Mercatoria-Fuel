import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Blueprint, render_template, request, redirect, session, send_file

from database import conectar
from utils.auth import requiere_login
from utils.constants import ROLES_ADMIN_PM

reportes_bp = Blueprint("reportes", __name__, url_prefix="/reportes")

_ROLES_REPORTE = ["admin", "pm", "supervisor"]


def _requiere_reporte():
    if "usuario" not in session:
        return redirect("/login")
    if session.get("rol") not in _ROLES_REPORTE:
        return redirect("/dashboard?access_error=Acceso+restringido+a+admin,+pm+y+supervisor")
    return None


def _header_row(ws, cols):
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")


# ── Índice ────────────────────────────────────────────────────────────────────

@reportes_bp.route("/")
def index():
    redir = requiere_login()
    if redir:
        return redir
    check = _requiere_reporte()
    if check:
        return check

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT id, nombre FROM gasolineras WHERE estado = 'activo' ORDER BY nombre")
    gasolineras = cur.fetchall()
    cur.execute("SELECT id, nombre, codigo FROM clientes WHERE activo = 1 ORDER BY nombre")
    clientes = cur.fetchall()
    conn.close()

    return render_template("reportes/index.html", gasolineras=gasolineras, clientes=clientes,
                           hoy=date.today().isoformat())


# ── Reporte Despachos ─────────────────────────────────────────────────────────

@reportes_bp.route("/despachos")
def despachos():
    redir = requiere_login()
    if redir:
        return redir
    check = _requiere_reporte()
    if check:
        return check

    filtro_cliente = request.args.get("cliente_id", "")
    filtro_gasolinera = request.args.get("gasolinera_id", "")
    filtro_desde = request.args.get("desde", "")
    filtro_hasta = request.args.get("hasta", "")

    conds = []
    params = []
    if filtro_cliente:
        conds.append("d.cliente_id = ?")
        params.append(filtro_cliente)
    if filtro_gasolinera:
        conds.append("d.gasolinera_id = ?")
        params.append(filtro_gasolinera)
    if filtro_desde:
        conds.append("d.fecha_despacho >= ?")
        params.append(filtro_desde)
    if filtro_hasta:
        conds.append("d.fecha_despacho <= ?")
        params.append(filtro_hasta)

    where = ("WHERE " + " AND ".join(conds)) if conds else ""

    conn = conectar()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT d.fecha_despacho, cl.nombre AS cliente, v.chapa,
               ch.nombre AS chofer, g.nombre AS gasolinera,
               t.numero_parcial AS tarjeta, d.litros_despachados,
               u.nombre AS operario
        FROM despachos d
        JOIN clientes cl ON cl.id = d.cliente_id
        JOIN vehiculos v ON v.id = d.unidad_id
        LEFT JOIN choferes ch ON ch.id = v.chofer_id
        JOIN gasolineras g ON g.id = d.gasolinera_id
        JOIN tarjetas t ON t.id = d.tarjeta_id
        JOIN usuarios u ON u.id = d.operario_id
        {where}
        ORDER BY d.fecha_despacho DESC
    """, params)
    filas = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Despachos"
    _header_row(ws, ["Fecha", "Cliente", "Chapa", "Chofer", "Gasolinera", "Tarjeta", "Litros", "Operario"])
    for f in filas:
        ws.append([f["fecha_despacho"], f["cliente"], f["chapa"],
                   f["chofer"] or "", f["gasolinera"],
                   f"**** {f['tarjeta']}", f["litros_despachados"], f["operario"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="reporte_despachos.xlsx")


# ── Reporte Conciliaciones ────────────────────────────────────────────────────

@reportes_bp.route("/conciliaciones")
def conciliaciones():
    redir = requiere_login()
    if redir:
        return redir
    check = _requiere_reporte()
    if check:
        return check

    filtro_gasolinera = request.args.get("gasolinera_id", "")
    filtro_mes = request.args.get("mes", "")

    conds = []
    params = []
    if filtro_gasolinera:
        conds.append("c.gasolinera_id = ?")
        params.append(filtro_gasolinera)
    if filtro_mes:
        conds.append("c.fecha LIKE ?")
        params.append(filtro_mes + "%")

    where = ("WHERE " + " AND ".join(conds)) if conds else ""

    conn = conectar()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT c.fecha, g.nombre AS gasolinera,
               c.saldo_fisico_inicio_l, c.saldo_fisico_fin_l,
               c.total_entrada_l, c.total_despachado_l,
               c.diferencia_l, c.estado
        FROM conciliaciones c
        JOIN gasolineras g ON g.id = c.gasolinera_id
        {where}
        ORDER BY c.fecha DESC
    """, params)
    filas = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliaciones"
    _header_row(ws, ["Fecha", "Gasolinera", "Saldo Inicio (L)", "Saldo Fin (L)",
                     "Entradas (L)", "Despachado (L)", "Diferencia (L)", "Estado"])
    for f in filas:
        ws.append([f["fecha"], f["gasolinera"], f["saldo_fisico_inicio_l"],
                   f["saldo_fisico_fin_l"], f["total_entrada_l"],
                   f["total_despachado_l"], f["diferencia_l"], f["estado"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="reporte_conciliaciones.xlsx")


# ── Reporte Consumo por Cliente ───────────────────────────────────────────────

@reportes_bp.route("/consumo")
def consumo():
    redir = requiere_login()
    if redir:
        return redir
    check = _requiere_reporte()
    if check:
        return check

    filtro_desde = request.args.get("desde", "")
    filtro_hasta = request.args.get("hasta", "")

    conds = []
    params = []
    if filtro_desde:
        conds.append("d.fecha_despacho >= ?")
        params.append(filtro_desde)
    if filtro_hasta:
        conds.append("d.fecha_despacho <= ?")
        params.append(filtro_hasta)

    where = ("WHERE " + " AND ".join(conds)) if conds else ""

    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT id, nombre FROM clientes WHERE activo = 1 ORDER BY nombre")
    clientes = cur.fetchall()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for cliente in clientes:
        ws = wb.create_sheet(title=cliente["nombre"][:31])

        # Por vehículo
        cur.execute(f"""
            SELECT v.chapa, COALESCE(SUM(d.litros_despachados), 0) AS total
            FROM despachos d
            JOIN vehiculos v ON v.id = d.unidad_id
            {where + (' AND ' if where else 'WHERE ')}d.cliente_id = ?
            GROUP BY v.chapa ORDER BY total DESC
        """, params + [cliente["id"]])
        por_vehiculo = cur.fetchall()

        ws.append([f"Cliente: {cliente['nombre']}"])
        ws["A1"].font = Font(bold=True, size=13)
        ws.append([])
        ws.append(["Consumo por Vehículo"])
        ws["A3"].font = Font(bold=True)
        _header_row(ws, ["Chapa", "Total Litros"])
        for r in por_vehiculo:
            ws.append([r["chapa"], r["total"]])

        ws.append([])
        ws.append(["Consumo por Chofer"])
        ws[f"A{ws.max_row}"].font = Font(bold=True)
        _header_row(ws, ["Chofer", "Total Litros"])
        cur.execute(f"""
            SELECT COALESCE(ch.nombre, 'Sin asignar') AS chofer,
                   COALESCE(SUM(d.litros_despachados), 0) AS total
            FROM despachos d
            JOIN vehiculos v ON v.id = d.unidad_id
            LEFT JOIN choferes ch ON ch.id = v.chofer_id
            {where + (' AND ' if where else 'WHERE ')}d.cliente_id = ?
            GROUP BY ch.nombre ORDER BY total DESC
        """, params + [cliente["id"]])
        for r in cur.fetchall():
            ws.append([r["chofer"], r["total"]])

    conn.close()

    if not wb.sheetnames:
        wb.create_sheet("Sin datos")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="reporte_consumo_clientes.xlsx")


# ── Reporte Tarjetas ──────────────────────────────────────────────────────────

@reportes_bp.route("/tarjetas")
def tarjetas():
    redir = requiere_login()
    if redir:
        return redir
    check = _requiere_reporte()
    if check:
        return check

    conn = conectar()
    cur = conn.cursor()
    cur.execute("""
        SELECT t.numero_parcial, g.nombre AS gasolinera,
               t.tipo_combustible, t.saldo_usable_l, t.saldo_retenido_l,
               t.estado,
               (SELECT MAX(r.fecha) FROM recargas_tarjetas r WHERE r.tarjeta_id = t.id) AS ultima_recarga
        FROM tarjetas t
        JOIN gasolineras g ON g.id = t.gasolinera_id
        ORDER BY g.nombre, t.numero_parcial
    """)
    filas = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tarjetas"
    _header_row(ws, ["Tarjeta", "Gasolinera", "Combustible", "Saldo Usable (L)",
                     "Saldo Retenido (L)", "Estado", "Última Recarga"])
    for f in filas:
        ws.append([f"**** {f['numero_parcial']}", f["gasolinera"], f["tipo_combustible"],
                   f["saldo_usable_l"], f["saldo_retenido_l"], f["estado"],
                   f["ultima_recarga"] or "—"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="reporte_tarjetas.xlsx")
