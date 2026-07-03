import io
from datetime import date

import openpyxl
from flask import Blueprint, render_template, request, redirect, session, send_file

from database import conectar
from utils.auth import requiere_login
from utils.constants import ROLES_ADMIN_PM

tl38_bp = Blueprint("tl38", __name__, url_prefix="/tl38")

_TIPOS_TL38 = ["entrada", "despacho", "ajuste"]


def _requiere_admin_pm():
    if session.get("rol") not in ROLES_ADMIN_PM:
        return redirect("/tl38/?access_error=Acceso+restringido+a+admin+y+pm")
    return None


# ── Dashboard ─────────────────────────────────────────────────────────────────

@tl38_bp.route("/")
def dashboard():
    redir = requiere_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()

    hoy = date.today()
    mes_inicio = hoy.replace(day=1).isoformat()
    if hoy.month == 12:
        mes_fin = hoy.replace(year=hoy.year + 1, month=1, day=1).isoformat()
    else:
        mes_fin = hoy.replace(month=hoy.month + 1, day=1).isoformat()

    cur.execute("""
        SELECT COALESCE(SUM(litros), 0) AS total FROM movimientos_tl38
        WHERE tipo = 'entrada' AND fecha >= ? AND fecha < ?
    """, (mes_inicio, mes_fin))
    total_entradas_mes = cur.fetchone()["total"] or 0

    cur.execute("""
        SELECT COALESCE(SUM(litros), 0) AS total FROM movimientos_tl38
        WHERE tipo = 'despacho' AND fecha >= ? AND fecha < ?
    """, (mes_inicio, mes_fin))
    total_despachos_mes = cur.fetchone()["total"] or 0

    diferencia_mes = total_entradas_mes - total_despachos_mes

    cur.execute("""
        SELECT m.*, g.nombre AS gasolinera_nombre, u.nombre AS responsable_nombre
        FROM movimientos_tl38 m
        LEFT JOIN gasolineras g ON g.id = m.gasolinera_id
        LEFT JOIN usuarios u ON u.id = m.responsable_id
        ORDER BY m.created_at DESC LIMIT 10
    """)
    recientes = cur.fetchall()

    cur.execute("SELECT id, nombre FROM gasolineras WHERE estado = 'activo' ORDER BY nombre")
    gasolineras = cur.fetchall()

    # 6-month chart data
    year, month = hoy.year, hoy.month - 5
    if month <= 0:
        month += 12
        year -= 1
    desde_6m = date(year, month, 1).isoformat()
    cur.execute("""
        SELECT SUBSTR(CAST(fecha AS TEXT), 1, 7) AS mes,
               COALESCE(SUM(CASE WHEN tipo = 'entrada' THEN litros ELSE 0 END), 0) AS entradas,
               COALESCE(SUM(CASE WHEN tipo = 'despacho' THEN litros ELSE 0 END), 0) AS despachos
        FROM movimientos_tl38
        WHERE fecha >= ?
        GROUP BY mes
        ORDER BY mes ASC
    """, (desde_6m,))
    chart_rows = cur.fetchall()
    chart_labels = [r["mes"] for r in chart_rows]
    chart_entradas = [round(float(r["entradas"]), 2) for r in chart_rows]
    chart_despachos = [round(float(r["despachos"]), 2) for r in chart_rows]

    conn.close()

    return render_template(
        "tl38/dashboard.html",
        total_entradas_mes=total_entradas_mes,
        total_despachos_mes=total_despachos_mes,
        diferencia_mes=diferencia_mes,
        recientes=recientes,
        gasolineras=gasolineras,
        tipos=_TIPOS_TL38,
        chart_labels=chart_labels,
        chart_entradas=chart_entradas,
        chart_despachos=chart_despachos,
    )


# ── Listado ───────────────────────────────────────────────────────────────────

@tl38_bp.route("/listado")
def listado():
    redir = requiere_login()
    if redir:
        return redir

    filtro_tipo = request.args.get("tipo", "")
    filtro_flota = request.args.get("flota", "")
    filtro_desde = request.args.get("desde", "")
    filtro_hasta = request.args.get("hasta", "")
    exportar = request.args.get("exportar", "")

    conn = conectar()
    cur = conn.cursor()

    condiciones = []
    params = []
    if filtro_tipo:
        condiciones.append("m.tipo = ?")
        params.append(filtro_tipo)
    if filtro_flota:
        condiciones.append("m.flota = ?")
        params.append(filtro_flota)
    if filtro_desde:
        condiciones.append("m.fecha >= ?")
        params.append(filtro_desde)
    if filtro_hasta:
        condiciones.append("m.fecha <= ?")
        params.append(filtro_hasta)

    where = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""

    cur.execute(f"""
        SELECT m.*, g.nombre AS gasolinera_nombre, u.nombre AS responsable_nombre
        FROM movimientos_tl38 m
        LEFT JOIN gasolineras g ON g.id = m.gasolinera_id
        LEFT JOIN usuarios u ON u.id = m.responsable_id
        {where}
        ORDER BY m.fecha DESC, m.created_at DESC
    """, params)
    movimientos = cur.fetchall()

    # Flotas disponibles para filtro
    cur.execute("SELECT DISTINCT flota FROM movimientos_tl38 WHERE flota IS NOT NULL ORDER BY flota")
    flotas = [r["flota"] for r in cur.fetchall()]

    conn.close()

    if exportar == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos TL38"
        ws.append(["Fecha", "Tipo", "Chapa", "Chofer", "Litros", "Tarjeta TL38", "Flota", "Gasolinera", "Observaciones", "Responsable"])
        for m in movimientos:
            ws.append([
                m["fecha"], m["tipo"], m["chapa"], m["chofer"],
                m["litros"], m["tarjeta_tl38"] or "", m["flota"],
                m["gasolinera_nombre"] or "", m["observaciones"] or "",
                m["responsable_nombre"]
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="tl38_movimientos.xlsx")

    return render_template(
        "tl38/listado.html",
        movimientos=movimientos,
        tipos=_TIPOS_TL38,
        flotas=flotas,
        filtro_tipo=filtro_tipo,
        filtro_flota=filtro_flota,
        filtro_desde=filtro_desde,
        filtro_hasta=filtro_hasta,
    )


# ── Detalle ───────────────────────────────────────────────────────────────────

@tl38_bp.route("/<int:id>")
def detalle(id):
    redir = requiere_login()
    if redir:
        return redir

    conn = conectar()
    cur = conn.cursor()
    cur.execute("""
        SELECT m.*, g.nombre AS gasolinera_nombre, u.nombre AS responsable_nombre
        FROM movimientos_tl38 m
        LEFT JOIN gasolineras g ON g.id = m.gasolinera_id
        LEFT JOIN usuarios u ON u.id = m.responsable_id
        WHERE m.id = ?
    """, (id,))
    mov = cur.fetchone()
    conn.close()

    if not mov:
        return redirect("/tl38/listado")

    return render_template("tl38/detalle.html", mov=mov)


# ── Crear movimiento ──────────────────────────────────────────────────────────

@tl38_bp.route("/crear", methods=["GET", "POST"])
def crear():
    redir = requiere_login()
    if redir:
        return redir
    check = _requiere_admin_pm()
    if check:
        return check

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT id, nombre FROM gasolineras WHERE estado = 'activo' ORDER BY nombre")
    gasolineras = cur.fetchall()

    if request.method == "POST":
        fecha = request.form.get("fecha", "").strip()
        tipo = request.form.get("tipo", "despacho").strip()
        chapa = request.form.get("chapa", "").strip().upper()
        chofer = request.form.get("chofer", "").strip()
        litros_str = request.form.get("litros", "0").strip()
        tarjeta_tl38 = request.form.get("tarjeta_tl38", "").strip() or None
        flota = request.form.get("flota", "599").strip() or "599"
        gasolinera_id = request.form.get("gasolinera_id", "").strip() or None
        observaciones = request.form.get("observaciones", "").strip() or None

        error = None
        if not fecha:
            error = "La fecha es obligatoria."
        elif not chapa:
            error = "La chapa es obligatoria."
        elif not chofer:
            error = "El nombre del chofer es obligatorio."
        else:
            try:
                litros = float(litros_str)
                if litros <= 0:
                    error = "Los litros deben ser mayores a 0."
            except ValueError:
                error = "Los litros deben ser un número válido."

        if error:
            conn.close()
            return render_template("tl38/crear.html", error=error, gasolineras=gasolineras,
                                   tipos=_TIPOS_TL38, hoy=date.today().isoformat())

        cur.execute("""
            INSERT INTO movimientos_tl38
                (fecha, gasolinera_id, tipo, chapa, chofer, litros, tarjeta_tl38, flota, observaciones, responsable_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (fecha, gasolinera_id, tipo, chapa, chofer, litros, tarjeta_tl38, flota, observaciones,
              session["user_id"]))
        conn.commit()
        conn.close()
        return redirect("/tl38/?ok=1")

    conn.close()
    return render_template("tl38/crear.html", error=None, gasolineras=gasolineras,
                           tipos=_TIPOS_TL38, hoy=date.today().isoformat())


# ── Importar Excel ────────────────────────────────────────────────────────────

@tl38_bp.route("/importar", methods=["GET", "POST"])
def importar():
    redir = requiere_login()
    if redir:
        return redir
    check = _requiere_admin_pm()
    if check:
        return check

    if request.method == "POST":
        archivo = request.files.get("archivo")
        if not archivo or not archivo.filename.endswith((".xlsx", ".xls")):
            return render_template("tl38/importar.html", error="Selecciona un archivo Excel (.xlsx o .xls).",
                                   resultado=None)

        conn = conectar()
        cur = conn.cursor()
        importados = 0
        fallidos = []

        _COL = {
            "fecha": ["fecha", "date"],
            "tipo": ["tipo", "type"],
            "chapa": ["chapa", "matricula", "placa"],
            "chofer": ["chofer", "conductor", "driver"],
            "litros": ["litros", "liters", "cantidad"],
            "tarjeta_tl38": ["tarjeta", "tarjeta_tl38", "card"],
            "flota": ["flota", "fleet"],
            "gasolinera": ["gasolinera", "station"],
            "observaciones": ["observaciones", "notas", "notes"],
        }

        try:
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
            headers_raw = [str(c.value).lower().strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            def col(field):
                for alias in _COL[field]:
                    if alias in headers_raw:
                        return headers_raw.index(alias)
                return None

            idx = {f: col(f) for f in _COL}

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    fecha = str(row[idx["fecha"]]).strip()[:10] if idx["fecha"] is not None and row[idx["fecha"]] else None
                    chapa = str(row[idx["chapa"]]).strip().upper() if idx["chapa"] is not None and row[idx["chapa"]] else None
                    chofer = str(row[idx["chofer"]]).strip() if idx["chofer"] is not None and row[idx["chofer"]] else None
                    litros_v = row[idx["litros"]] if idx["litros"] is not None else None
                    litros = float(str(litros_v).replace(",", ".")) if litros_v else None

                    if not fecha or not chapa or not chofer or not litros:
                        fallidos.append(f"Fila {i}: datos incompletos")
                        continue

                    tipo = str(row[idx["tipo"]]).strip().lower() if idx["tipo"] is not None and row[idx["tipo"]] else "despacho"
                    if tipo not in _TIPOS_TL38:
                        tipo = "despacho"
                    tarjeta_tl38 = str(row[idx["tarjeta_tl38"]]).strip() if idx["tarjeta_tl38"] is not None and row[idx["tarjeta_tl38"]] else None
                    flota = str(row[idx["flota"]]).strip() if idx["flota"] is not None and row[idx["flota"]] else "599"
                    observaciones = str(row[idx["observaciones"]]).strip() if idx["observaciones"] is not None and row[idx["observaciones"]] else None

                    gasolinera_id = None
                    if idx["gasolinera"] is not None and row[idx["gasolinera"]]:
                        g_nombre = str(row[idx["gasolinera"]]).strip()
                        cur.execute("SELECT id FROM gasolineras WHERE nombre = ?", (g_nombre,))
                        gr = cur.fetchone()
                        if gr:
                            gasolinera_id = gr["id"]

                    cur.execute("""
                        INSERT INTO movimientos_tl38
                            (fecha, gasolinera_id, tipo, chapa, chofer, litros, tarjeta_tl38, flota, observaciones, responsable_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (fecha, gasolinera_id, tipo, chapa, chofer, litros, tarjeta_tl38, flota, observaciones, session["user_id"]))
                    importados += 1
                except Exception as e:
                    fallidos.append(f"Fila {i}: {e}")

            conn.commit()
        except Exception as ex:
            conn.close()
            return render_template("tl38/importar.html", error=f"Error al leer el archivo: {ex}", resultado=None)

        conn.close()
        return render_template("tl38/importar.html", error=None,
                               resultado={"importados": importados, "fallidos": fallidos})

    return render_template("tl38/importar.html", error=None, resultado=None)
