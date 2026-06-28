import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, session
from database import conectar
from utils.constants import ROLES_ADMIN_PM
from utils.auth import requiere_login

choferes_bp = Blueprint("choferes", __name__, url_prefix="/choferes")

_COLUMN_MAP = {
    "nombre":                 "nombre",
    "ci":                     "ci",
    "carnet":                 "ci",
    "carnet de identidad":    "ci",
    "licencia_numero":        "licencia_numero",
    "licencia número":       "licencia_numero",
    "licencia numero":        "licencia_numero",
    "n° licencia":           "licencia_numero",
    "licencia":               "licencia_numero",
    "licencia_vencimiento":   "licencia_vencimiento",
    "vencimiento":            "licencia_vencimiento",
    "vencimiento licencia":   "licencia_vencimiento",
    "vence":                  "licencia_vencimiento",
    "telefono":               "telefono",
    "teléfono":              "telefono",
    "observaciones":          "observaciones",
    "cliente":                "cliente_codigo",
    "cliente_codigo":         "cliente_codigo",
    "codigo cliente":         "cliente_codigo",
    "código cliente":        "cliente_codigo",
}


def _parse_fecha(val):
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _requiere_admin_pm():
    return session.get("rol") not in ROLES_ADMIN_PM


def _registrar_auditoria(usuario_id, accion, tabla, registro_id, valor_anterior=None, valor_nuevo=None):
    try:
        conn = conectar()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO auditoria
                (usuario_id, accion, tabla_afectada, registro_id, valor_anterior, valor_nuevo, ip, user_agent)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            usuario_id, accion, tabla, registro_id,
            str(valor_anterior) if valor_anterior else None,
            str(valor_nuevo) if valor_nuevo else None,
            request.remote_addr,
            request.headers.get("User-Agent", "")[:512],
        ))
        conn.commit()
        conn.close()
    except Exception:
        from flask import current_app
        current_app.logger.exception("Error registrando auditoría")


def _importar_excel(archivo):
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(archivo.read()), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(cell.value or "").strip().lower() for cell in header_row]

    col_indices = {}
    for i, h in enumerate(headers):
        mapped = _COLUMN_MAP.get(h)
        if mapped and mapped not in col_indices:
            col_indices[mapped] = i

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT id, codigo FROM clientes WHERE activo = 1")
    clientes_map = {row["codigo"]: row["id"] for row in cur.fetchall()}

    importados = 0
    fallos = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        def get(col):
            idx = col_indices.get(col)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        nombre = get("nombre")
        if not nombre:
            fallos.append(f"Fila {row_num}: nombre vacío — fila omitida")
            continue

        ci = get("ci")
        if not ci:
            fallos.append(f"Fila {row_num} ({nombre}): C.I. vacío — fila omitida")
            continue

        cliente_codigo = get("cliente_codigo")
        cliente_id = clientes_map.get(cliente_codigo) if cliente_codigo else None
        if not cliente_id:
            fallos.append(f"Fila {row_num} ({nombre}): cliente '{cliente_codigo}' no encontrado")
            continue

        licencia_numero = get("licencia_numero")

        venc_raw = row[col_indices["licencia_vencimiento"]] if "licencia_vencimiento" in col_indices else None
        licencia_vencimiento = _parse_fecha(venc_raw)

        telefono = get("telefono")
        observaciones = get("observaciones")

        try:
            cur.execute("SELECT id FROM choferes WHERE ci = ?", (ci,))
            existente = cur.fetchone()
            if existente:
                cur.execute("""
                    UPDATE choferes
                    SET cliente_id = ?, nombre = ?, licencia_numero = ?,
                        licencia_vencimiento = ?, telefono = ?, observaciones = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE ci = ?
                """, (cliente_id, nombre, licencia_numero, licencia_vencimiento,
                      telefono, observaciones, ci))
            else:
                cur.execute("""
                    INSERT INTO choferes
                        (cliente_id, nombre, ci, licencia_numero, licencia_vencimiento,
                         telefono, observaciones)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (cliente_id, nombre, ci, licencia_numero, licencia_vencimiento,
                      telefono, observaciones))
            importados += 1
        except Exception as exc:
            fallos.append(f"Fila {row_num} ({nombre}): error al guardar — {exc}")

    conn.commit()
    conn.close()
    return importados, fallos


@choferes_bp.route("/")
def listado():
    return redirect("/unidades")


@choferes_bp.route("/legacy_listado")
def _listado_legacy():
    redir = requiere_login()
    if redir:
        return redir

    buscar = request.args.get("buscar", "").strip()
    filtro_cliente = request.args.get("cliente_id", "").strip()
    filtro_estado = request.args.get("estado", "").strip()

    condiciones = []
    params = []

    if buscar:
        condiciones.append("(ch.nombre LIKE ? OR ch.ci LIKE ? OR ch.licencia_numero LIKE ?)")
        like = f"%{buscar}%"
        params.extend([like, like, like])
    if filtro_cliente:
        condiciones.append("ch.cliente_id = ?")
        params.append(filtro_cliente)
    if filtro_estado:
        condiciones.append("ch.estado = ?")
        params.append(filtro_estado)

    where = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""

    conn = conectar()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT ch.id, ch.cliente_id, ch.nombre, ch.ci, ch.licencia_numero, ch.licencia_vencimiento,
               ch.telefono, ch.estado,
               c.nombre AS cliente_nombre, c.codigo AS cliente_codigo
        FROM choferes ch
        JOIN clientes c ON c.id = ch.cliente_id
        {where}
        ORDER BY ch.nombre ASC
    """, params)
    lista = cur.fetchall()

    cur.execute("SELECT id, nombre, codigo FROM clientes WHERE activo = 1 ORDER BY nombre ASC")
    clientes = cur.fetchall()
    conn.close()

    hoy = datetime.date.today().isoformat()
    limite_30 = (datetime.date.today() + datetime.timedelta(days=30)).isoformat()

    return render_template(
        "choferes/listado.html",
        lista=lista,
        clientes=clientes,
        buscar=buscar,
        filtro_cliente=filtro_cliente,
        filtro_estado=filtro_estado,
        hoy=hoy,
        limite_30=limite_30,
    )


@choferes_bp.route("/crear", methods=["GET", "POST"])
def crear():
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/choferes?access_error=Solo+Admin+y+PM+pueden+crear+choferes")

    error = None

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, codigo FROM clientes WHERE activo = 1 ORDER BY nombre ASC")
    clientes = cur.fetchall()
    conn.close()

    if request.method == "POST":
        cliente_id = request.form.get("cliente_id", "").strip()
        nombre = request.form.get("nombre", "").strip()
        ci = request.form.get("ci", "").strip()
        licencia_numero = request.form.get("licencia_numero", "").strip()
        licencia_vencimiento = request.form.get("licencia_vencimiento", "").strip() or None
        telefono = request.form.get("telefono", "").strip()
        observaciones = request.form.get("observaciones", "").strip()
        estado = request.form.get("estado", "activo").strip()

        if not nombre:
            error = "El nombre es obligatorio."
        elif not ci:
            error = "El C.I. es obligatorio."
        elif not cliente_id:
            error = "Debe seleccionar un cliente."
        else:
            conn = conectar()
            cur = conn.cursor()
            cur.execute("SELECT id FROM choferes WHERE ci = ?", (ci,))
            if cur.fetchone():
                error = f"Ya existe un chofer con el C.I. {ci}."
                conn.close()
            else:
                cur.execute("""
                    INSERT INTO choferes
                        (cliente_id, nombre, ci, licencia_numero, licencia_vencimiento,
                         telefono, observaciones, estado)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (cliente_id, nombre, ci, licencia_numero, licencia_vencimiento,
                      telefono, observaciones, estado))
                nuevo_id = cur.lastrowid
                conn.commit()
                conn.close()
                _registrar_auditoria(
                    session.get("user_id"), "Creó chofer", "choferes", nuevo_id,
                    valor_nuevo={"nombre": nombre, "ci": ci, "cliente_id": cliente_id}
                )
                return redirect("/choferes?ok=1")

    return render_template(
        "choferes/crear.html",
        error=error,
        clientes=clientes,
        cliente_preseleccionado=request.args.get("cliente_id", ""),
    )


@choferes_bp.route("/<int:id>/editar", methods=["GET", "POST"])
def editar(id):
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/choferes?access_error=Solo+Admin+y+PM+pueden+editar+choferes")

    conn = conectar()
    cur = conn.cursor()
    error = None

    cur.execute("SELECT id, nombre, codigo FROM clientes WHERE activo = 1 ORDER BY nombre ASC")
    clientes = cur.fetchall()

    if request.method == "POST":
        cliente_id = request.form.get("cliente_id", "").strip()
        nombre = request.form.get("nombre", "").strip()
        ci = request.form.get("ci", "").strip()
        licencia_numero = request.form.get("licencia_numero", "").strip()
        licencia_vencimiento = request.form.get("licencia_vencimiento", "").strip() or None
        telefono = request.form.get("telefono", "").strip()
        observaciones = request.form.get("observaciones", "").strip()
        estado = request.form.get("estado", "activo").strip()

        if not nombre:
            error = "El nombre es obligatorio."
        elif not ci:
            error = "El C.I. es obligatorio."
        elif not cliente_id:
            error = "Debe seleccionar un cliente."
        else:
            cur.execute("SELECT id FROM choferes WHERE ci = ? AND id != ?", (ci, id))
            if cur.fetchone():
                error = f"Ya existe otro chofer con el C.I. {ci}."
            else:
                cur.execute("SELECT * FROM choferes WHERE id = ?", (id,))
                anterior = dict(cur.fetchone() or {})
                cur.execute("""
                    UPDATE choferes
                    SET cliente_id = ?, nombre = ?, ci = ?, licencia_numero = ?,
                        licencia_vencimiento = ?, telefono = ?, observaciones = ?,
                        estado = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (cliente_id, nombre, ci, licencia_numero, licencia_vencimiento,
                      telefono, observaciones, estado, id))
                conn.commit()
                conn.close()
                _registrar_auditoria(
                    session.get("user_id"), "Editó chofer", "choferes", id,
                    valor_anterior=anterior,
                    valor_nuevo={"nombre": nombre, "ci": ci, "estado": estado}
                )
                return redirect("/choferes?ok=1")

    cur.execute("SELECT * FROM choferes WHERE id = ?", (id,))
    chofer = cur.fetchone()
    conn.close()

    if not chofer:
        return redirect("/choferes")

    return render_template(
        "choferes/editar.html",
        chofer=chofer,
        clientes=clientes,
        error=error,
    )


@choferes_bp.route("/<int:id>/toggle", methods=["POST"])
def toggle_estado(id):
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/choferes?access_error=Sin+permisos")

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT estado FROM choferes WHERE id = ?", (id,))
    row = cur.fetchone()
    if row:
        nuevo_estado = "inactivo" if row["estado"] == "activo" else "activo"
        cur.execute(
            "UPDATE choferes SET estado = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (nuevo_estado, id)
        )
        conn.commit()
        _registrar_auditoria(
            session.get("user_id"),
            f"Cambió estado a {nuevo_estado}", "choferes", id,
            valor_anterior={"estado": row["estado"]},
            valor_nuevo={"estado": nuevo_estado}
        )
    conn.close()
    return redirect("/choferes")


@choferes_bp.route("/importar", methods=["GET", "POST"])
def importar():
    return redirect("/unidades/importar")


@choferes_bp.route("/legacy_importar", methods=["GET", "POST"])
def _importar_legacy():
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/choferes?access_error=Solo+Admin+y+PM+pueden+importar+choferes")

    resultado = None

    if request.method == "POST":
        archivo = request.files.get("archivo")
        if not archivo or not archivo.filename.endswith(".xlsx"):
            resultado = {"error": "Debe seleccionar un archivo .xlsx válido."}
        else:
            try:
                importados, fallos = _importar_excel(archivo)
                resultado = {"importados": importados, "fallos": fallos}
            except Exception as exc:
                resultado = {"error": f"Error al procesar el archivo: {exc}"}

    return render_template("choferes/importar.html", resultado=resultado)
