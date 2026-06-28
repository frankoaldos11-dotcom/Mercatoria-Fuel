import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, session
from database import conectar
from utils.constants import TIPOS_COMBUSTIBLE, TIPOS_COMBUSTIBLE_LABELS, ROLES_ADMIN_PM
from utils.auth import requiere_login

vehiculos_bp = Blueprint("vehiculos", __name__, url_prefix="/vehiculos")

_COMBUSTIBLE_NORM = {
    "diesel":             "diesel",
    "diésel":            "diesel",
    "gasolina regular":   "gasolina_regular",
    "regular":            "gasolina_regular",
    "gasolina_regular":   "gasolina_regular",
    "gasolina especial":  "gasolina_especial",
    "especial":           "gasolina_especial",
    "gasolina_especial":  "gasolina_especial",
}

_COLUMN_MAP = {
    "chapa":              "chapa",
    "matrícula":         "chapa",
    "matricula":          "chapa",
    "marca":              "marca",
    "modelo":             "modelo",
    "año":               "anio",
    "año fabricación":   "anio",
    "anio":               "anio",
    "ano":                "anio",
    "tipo_combustible":   "tipo_combustible",
    "tipo combustible":   "tipo_combustible",
    "combustible":        "tipo_combustible",
    "cuota_mensual_l":    "cuota_mensual_l",
    "cuota mensual":      "cuota_mensual_l",
    "cuota mensual (l)":  "cuota_mensual_l",
    "cuota (l)":          "cuota_mensual_l",
    "color":              "color",
    "observaciones":      "observaciones",
    "cliente":            "cliente_codigo",
    "cliente_codigo":     "cliente_codigo",
    "codigo cliente":     "cliente_codigo",
    "código cliente":    "cliente_codigo",
}


def _requiere_admin_pm():
    return session.get("rol") not in ROLES_ADMIN_PM


def _registrar_auditoria(usuario_id, accion, tabla, registro_id, valor_anterior=None, valor_nuevo=None):
    try:
        conn = conectar()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO auditoria
                (usuario_id, accion, tabla_afectada, registro_id, valor_anterior, valor_nuevo, ip, user_agent)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            usuario_id, accion, tabla, registro_id,
            str(valor_anterior) if valor_anterior else None,
            str(valor_nuevo) if valor_nuevo else None,
            request.remote_addr,
            request.headers.get("User-Agent", "")[:512],
        ))
        conn.commit()
        conn.close()
    except Exception:
        from flask import current_app
        current_app.logger.exception("Error registrando auditoría")


def _importar_excel(archivo):
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(archivo.read()), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(cell.value or "").strip().lower() for cell in header_row]

    col_indices = {}
    for i, h in enumerate(headers):
        mapped = _COLUMN_MAP.get(h)
        if mapped and mapped not in col_indices:
            col_indices[mapped] = i

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT id, codigo FROM clientes WHERE activo = 1")
    clientes_map = {row["codigo"]: row["id"] for row in cur.fetchall()}

    importados = 0
    fallos = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        def get(col):
            idx = col_indices.get(col)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        chapa = get("chapa")
        if not chapa:
            fallos.append(f"Fila {row_num}: chapa vacía — fila omitida")
            continue
        chapa = chapa.upper()

        cliente_codigo = get("cliente_codigo")
        cliente_id = clientes_map.get(cliente_codigo) if cliente_codigo else None
        if not cliente_id:
            fallos.append(f"Fila {row_num} ({chapa}): cliente '{cliente_codigo}' no encontrado")
            continue

        combustible_raw = (get("tipo_combustible") or "").strip().lower()
        tipo_combustible = _COMBUSTIBLE_NORM.get(combustible_raw)
        if not tipo_combustible:
            fallos.append(f"Fila {row_num} ({chapa}): tipo de combustible inválido '{combustible_raw}'")
            continue

        marca = get("marca")
        modelo = get("modelo")
        anio_raw = get("anio")
        try:
            anio = int(float(anio_raw)) if anio_raw else None
        except (ValueError, TypeError):
            anio = None

        cuota_raw = get("cuota_mensual_l")
        try:
            cuota = float(cuota_raw) if cuota_raw else None
        except (ValueError, TypeError):
            cuota = None

        color = get("color")
        observaciones = get("observaciones")

        try:
            cur.execute("SELECT id FROM vehiculos WHERE chapa = ?", (chapa,))
            existente = cur.fetchone()
            if existente:
                cur.execute("""
                    UPDATE vehiculos
                    SET cliente_id = ?, marca = ?, modelo = ?, anio = ?,
                        tipo_combustible = ?, cuota_mensual_l = ?, color = ?,
                        observaciones = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE chapa = ?
                """, (cliente_id, marca, modelo, anio, tipo_combustible, cuota, color, observaciones, chapa))
            else:
                cur.execute("""
                    INSERT INTO vehiculos
                        (cliente_id, chapa, marca, modelo, anio, tipo_combustible,
                         cuota_mensual_l, color, observaciones)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (cliente_id, chapa, marca, modelo, anio, tipo_combustible, cuota, color, observaciones))
            importados += 1
        except Exception as exc:
            fallos.append(f"Fila {row_num} ({chapa}): error al guardar — {exc}")

    conn.commit()
    conn.close()
    return importados, fallos


@vehiculos_bp.route("/")
def listado():
    return redirect("/unidades")


@vehiculos_bp.route("/legacy_listado")
def _listado_legacy():
    redir = requiere_login()
    if redir:
        return redir

    buscar = request.args.get("buscar", "").strip()
    filtro_cliente = request.args.get("cliente_id", "").strip()
    filtro_combustible = request.args.get("combustible", "").strip()
    filtro_estado = request.args.get("estado", "").strip()

    condiciones = []
    params = []

    if buscar:
        condiciones.append("(v.chapa LIKE ? OR v.marca LIKE ? OR v.modelo LIKE ?)")
        like = f"%{buscar}%"
        params.extend([like, like, like])
    if filtro_cliente:
        condiciones.append("v.cliente_id = ?")
        params.append(filtro_cliente)
    if filtro_combustible:
        condiciones.append("v.tipo_combustible = ?")
        params.append(filtro_combustible)
    if filtro_estado:
        condiciones.append("v.estado = ?")
        params.append(filtro_estado)

    where = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""

    conn = conectar()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT v.id, v.cliente_id, v.chapa, v.marca, v.modelo, v.anio, v.tipo_combustible,
               v.cuota_mensual_l, v.color, v.estado,
               c.nombre AS cliente_nombre, c.codigo AS cliente_codigo
        FROM vehiculos v
        JOIN clientes c ON c.id = v.cliente_id
        {where}
        ORDER BY v.chapa ASC
    """, params)
    lista = cur.fetchall()

    cur.execute("SELECT id, nombre, codigo FROM clientes WHERE activo = 1 ORDER BY nombre ASC")
    clientes = cur.fetchall()
    conn.close()

    return render_template(
        "vehiculos/listado.html",
        lista=lista,
        clientes=clientes,
        buscar=buscar,
        filtro_cliente=filtro_cliente,
        filtro_combustible=filtro_combustible,
        filtro_estado=filtro_estado,
        tipos_combustible=TIPOS_COMBUSTIBLE,
        combustible_labels=TIPOS_COMBUSTIBLE_LABELS,
    )


@vehiculos_bp.route("/crear", methods=["GET", "POST"])
def crear():
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/vehiculos?access_error=Solo+Admin+y+PM+pueden+crear+vehículos")

    error = None

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, codigo FROM clientes WHERE activo = 1 ORDER BY nombre ASC")
    clientes = cur.fetchall()
    conn.close()

    if request.method == "POST":
        cliente_id = request.form.get("cliente_id", "").strip()
        chapa = request.form.get("chapa", "").strip().upper()
        marca = request.form.get("marca", "").strip()
        modelo = request.form.get("modelo", "").strip()
        anio_str = request.form.get("anio", "").strip()
        tipo_combustible = request.form.get("tipo_combustible", "").strip()
        cuota_str = request.form.get("cuota_mensual_l", "").strip()
        color = request.form.get("color", "").strip()
        observaciones = request.form.get("observaciones", "").strip()
        estado = request.form.get("estado", "activo").strip()

        if not chapa:
            error = "La chapa es obligatoria."
        elif not cliente_id:
            error = "Debe seleccionar un cliente."
        elif tipo_combustible not in TIPOS_COMBUSTIBLE:
            error = "Tipo de combustible no válido."
        else:
            try:
                anio = int(anio_str) if anio_str else None
            except ValueError:
                anio = None
            try:
                cuota = float(cuota_str) if cuota_str else None
            except ValueError:
                cuota = None

            conn = conectar()
            cur = conn.cursor()
            cur.execute("SELECT id FROM vehiculos WHERE chapa = ?", (chapa,))
            if cur.fetchone():
                error = f"Ya existe un vehículo con la chapa {chapa}."
                conn.close()
            else:
                cur.execute("""
                    INSERT INTO vehiculos
                        (cliente_id, chapa, marca, modelo, anio, tipo_combustible,
                         cuota_mensual_l, color, observaciones, estado)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (cliente_id, chapa, marca, modelo, anio, tipo_combustible,
                      cuota, color, observaciones, estado))
                nuevo_id = cur.lastrowid
                conn.commit()
                conn.close()
                _registrar_auditoria(
                    session.get("user_id"), "Creó vehículo", "vehiculos", nuevo_id,
                    valor_nuevo={"chapa": chapa, "cliente_id": cliente_id, "tipo_combustible": tipo_combustible}
                )
                return redirect("/vehiculos?ok=1")

    return render_template(
        "vehiculos/crear.html",
        error=error,
        clientes=clientes,
        tipos_combustible=TIPOS_COMBUSTIBLE,
        combustible_labels=TIPOS_COMBUSTIBLE_LABELS,
        cliente_preseleccionado=request.args.get("cliente_id", ""),
    )


@vehiculos_bp.route("/<int:id>/editar", methods=["GET", "POST"])
def editar(id):
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/vehiculos?access_error=Solo+Admin+y+PM+pueden+editar+vehículos")

    conn = conectar()
    cur = conn.cursor()
    error = None

    cur.execute("SELECT id, nombre, codigo FROM clientes WHERE activo = 1 ORDER BY nombre ASC")
    clientes = cur.fetchall()

    if request.method == "POST":
        cliente_id = request.form.get("cliente_id", "").strip()
        chapa = request.form.get("chapa", "").strip().upper()
        marca = request.form.get("marca", "").strip()
        modelo = request.form.get("modelo", "").strip()
        anio_str = request.form.get("anio", "").strip()
        tipo_combustible = request.form.get("tipo_combustible", "").strip()
        cuota_str = request.form.get("cuota_mensual_l", "").strip()
        color = request.form.get("color", "").strip()
        observaciones = request.form.get("observaciones", "").strip()
        estado = request.form.get("estado", "activo").strip()

        if not chapa:
            error = "La chapa es obligatoria."
        elif not cliente_id:
            error = "Debe seleccionar un cliente."
        elif tipo_combustible not in TIPOS_COMBUSTIBLE:
            error = "Tipo de combustible no válido."
        else:
            try:
                anio = int(anio_str) if anio_str else None
            except ValueError:
                anio = None
            try:
                cuota = float(cuota_str) if cuota_str else None
            except ValueError:
                cuota = None

            cur.execute("SELECT id FROM vehiculos WHERE chapa = ? AND id != ?", (chapa, id))
            if cur.fetchone():
                error = f"Ya existe otro vehículo con la chapa {chapa}."
            else:
                cur.execute("SELECT * FROM vehiculos WHERE id = ?", (id,))
                anterior = dict(cur.fetchone() or {})
                cur.execute("""
                    UPDATE vehiculos
                    SET cliente_id = ?, chapa = ?, marca = ?, modelo = ?, anio = ?,
                        tipo_combustible = ?, cuota_mensual_l = ?, color = ?,
                        observaciones = ?, estado = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (cliente_id, chapa, marca, modelo, anio, tipo_combustible,
                      cuota, color, observaciones, estado, id))
                conn.commit()
                conn.close()
                _registrar_auditoria(
                    session.get("user_id"), "Editó vehículo", "vehiculos", id,
                    valor_anterior=anterior,
                    valor_nuevo={"chapa": chapa, "tipo_combustible": tipo_combustible, "estado": estado}
                )
                return redirect("/vehiculos?ok=1")

    cur.execute("SELECT * FROM vehiculos WHERE id = ?", (id,))
    vehiculo = cur.fetchone()
    conn.close()

    if not vehiculo:
        return redirect("/vehiculos")

    return render_template(
        "vehiculos/editar.html",
        vehiculo=vehiculo,
        clientes=clientes,
        error=error,
        tipos_combustible=TIPOS_COMBUSTIBLE,
        combustible_labels=TIPOS_COMBUSTIBLE_LABELS,
    )


@vehiculos_bp.route("/<int:id>/toggle", methods=["POST"])
def toggle_estado(id):
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/vehiculos?access_error=Sin+permisos")

    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT estado FROM vehiculos WHERE id = ?", (id,))
    row = cur.fetchone()
    if row:
        nuevo_estado = "inactivo" if row["estado"] == "activo" else "activo"
        cur.execute(
            "UPDATE vehiculos SET estado = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (nuevo_estado, id)
        )
        conn.commit()
        _registrar_auditoria(
            session.get("user_id"),
            f"Cambió estado a {nuevo_estado}", "vehiculos", id,
            valor_anterior={"estado": row["estado"]},
            valor_nuevo={"estado": nuevo_estado}
        )
    conn.close()
    return redirect("/vehiculos")


@vehiculos_bp.route("/importar", methods=["GET", "POST"])
def importar():
    return redirect("/unidades/importar")


@vehiculos_bp.route("/legacy_importar", methods=["GET", "POST"])
def _importar_legacy():
    redir = requiere_login()
    if redir:
        return redir
    if _requiere_admin_pm():
        return redirect("/vehiculos?access_error=Solo+Admin+y+PM+pueden+importar+vehículos")

    resultado = None

    if request.method == "POST":
        archivo = request.files.get("archivo")
        if not archivo or not archivo.filename.endswith(".xlsx"):
            resultado = {"error": "Debe seleccionar un archivo .xlsx válido."}
        else:
            try:
                importados, fallos = _importar_excel(archivo)
                resultado = {"importados": importados, "fallos": fallos}
            except Exception as exc:
                resultado = {"error": f"Error al procesar el archivo: {exc}"}

    return render_template("vehiculos/importar.html", resultado=resultado)
